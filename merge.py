from openpyxl import load_workbook

main_workbook = load_workbook("data/result.xlsx")
result_sheet = main_workbook["Sheet1"]


# function that handles unmerging cells. We need to unmerge all merged cells and then give the ones that are empty a
# value because all cells that will contain data need to have a value.
def handle_merged_cells():
    merged_cells = result_sheet.merged_cells.ranges
    merged_cells_list = [item.coord for item in merged_cells]
    for merge_cell in merged_cells_list:
        left_cell_value = result_sheet[merge_cell.split(":")[0]].value
        result_sheet.unmerge_cells(merge_cell)
        result_sheet[merge_cell.split(":")[1]] = left_cell_value


# This function is mainly used for the header. If the main header doesn't have a % sign, then
# we want to merge the 2 cells (n, %) adjacent to one another and only keep the count.
def combine_cells(cell_list):
    count = 0
    cell_coordinate = []
    while count < len(cell_list) - 1:
        cell_group = f"{cell_list[count].coordinate}:{cell_list[count + 1].coordinate}"
        cell_coordinate.append(cell_group)
        count += 2
    return cell_coordinate


# in this function, we are trying to fetch the value of the adjacent cell(ie the label) of the current cell we are on.
# we used this mainly for the header row that we were trying to merge.

def fetch_label(current_cell):
    adjacent_cell_value = updated_ws[current_cell].offset(row=0, column=-1).value
    return adjacent_cell_value


def split_cells(cell_value):
    f_cell = updated_ws[cell_value.split(":")[0]]
    s_cell = updated_ws[cell_value.split(":")[1]]
    f_value = f_cell.value
    s_value = s_cell.value
    return [f_cell, s_cell, f_value, s_value]


# unmerge all the merged cells in the sheet and put new values in the empty right cell
handle_merged_cells()

# save the new workbook so you can work with it.
main_workbook.save("data/new_results.xlsx")

# load the new workbook(ie the result file) and the sheet you want to work on.
updated_wb = load_workbook("data/new_results.xlsx", data_only=True)
updated_ws = updated_wb["Sheet1"]

# load the data workbook and the sheet you want to work on.
data_wb = load_workbook("data/test.xlsx", data_only=True)
data_ws = data_wb.active

# we want to loop through the result file and the data file at the same time. but we need to keep track of what row
# of data we pasted from data file so we can continue from there on the next iteration. so we define a counter called
# i. the first loop loops from the third row and second column of the result file. the row variable is a tuple of all
# the cells on the current row it is looping. if any of the value in that row is empty, then we move to the next row
# without doing anything. the second for loop loops through the second row and second column of the data file. we
# fetch the index and the value to check what row we are looping on. if the current index matches the counter we kept
# track of, we iterate through the row(the cells of the result file) and update each cell to be the value of the data
# at the index of the data file. we finish that and update i and then break out of the data loop. This for loop
# basically pastes the data from the data file to the result file.
i = 0
for row in updated_ws.iter_rows(min_row=3, min_col=2):
    row_values = tuple(row_cell.value for row_cell in row)
    if any(map(lambda x: x is None, row_values)):
        continue
    for index2, row2 in enumerate(data_ws.iter_rows(min_row=2, min_col=2)):
        if index2 == i:
            for index, cell in enumerate(row):
                updated_ws[cell.coordinate] = row2[index].value
            i += 1
            break


# This for loop loops through the result file from row 3 and column 2. if any of the cells in a row is empty,
# we continue to the next row. We basically want to format our values to look the way that we want it to look. so we
# get the label of the first cell of the current row. Then we combine the 2 cells representing data for a column(ie
# n, %) Then we loop through the grouped cell. but then we unpack each grouped cell values in separate variables. if
# the label of the current row has a mean, then we merge the cell that we just unmerged and then format the value of
# the merged cell to how we want it. we do the same for 95%, median, and otherwise just have n and % in separate
# cells. Then we save the excel sheet.

for row in updated_ws.iter_rows(min_row=3, min_col=2):
    row_values = tuple(row_cell.value for row_cell in row)
    if any(map(lambda x: x is None, row_values)):
        continue

    first_cell = row[0].coordinate
    label_value = fetch_label(first_cell)
    grouped_cells = combine_cells(row)

    for cell in grouped_cells:
        first_cell, second_cell, first_value, second_value = split_cells(cell)
        if "mean" in label_value.lower():
            updated_ws.merge_cells(cell)
            updated_ws[cell.split(":")[0]] = f"{first_value}({second_value})"
        elif any(value in label_value.lower() for value in ("95% ci", "q1, q3", "minimum, maximum", "min, max")):
            updated_ws.merge_cells(cell)
            updated_ws[cell.split(":")[0]] = f"({first_value}, {second_value})"
        elif any(value in label_value.lower() for value in ("median", "total patient sample", "number of patients")):
            updated_ws.merge_cells(cell)
            updated_ws[cell.split(":")[0]] = f"{first_value}"
        elif any(value in label_value.lower() for value in ("12", "24", "36", "48", "60")):
            updated_ws[cell.split(":")[0]] = f"{first_value}%"
            updated_ws[cell.split(":")[1]] = f"{second_value}"
        else:
            updated_ws[cell.split(":")[0]] = f"{first_value}"
            updated_ws[cell.split(":")[1]] = f"{second_value}%"

updated_wb.save("data/new_results.xlsx")