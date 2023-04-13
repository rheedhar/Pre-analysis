from openpyxl import load_workbook
import sys


def extract_sheets():
    # Collect input data from user
    system_arguments = sys.argv

    # if cl arguments is less than 2, exit from the entire program
    if len(system_arguments) != 5:
        sys.exit("Please ensure to enter the correct number of file names")

    # save cl arguments in seperate variables
    name_workbook, name_sheet, output_table, data_file = system_arguments[1:]

    # if any of the variables is empty, exit from the entire program
    if " " in (name_workbook, name_sheet, output_table, data_file):
        sys.exit("Please ensure to enter a file name")

    # load the table shells workbook and extract all the sheet names
    main_workbook = load_workbook(f"data/{name_workbook}")
    sheets = main_workbook.sheetnames

    # if more than one sheet in wb, keep only the sheet name that is a match
    if len(sheets) > 1:
        for s in sheets:
            if s != name_sheet:
                sheet_name = main_workbook.get_sheet_by_name(s)
                main_workbook.remove_sheet(sheet_name)

    # save the workbook with only that sheet, and the sheet in a new variable.
    main_workbook.save(f"data/{name_workbook}")
    result_sheet = main_workbook.active

    # load the data workbook and save active sheet in its variable.
    data_workbook = load_workbook(f"data/{data_file}")
    data_sheet = data_workbook.active

    # return all documents
    return [main_workbook, result_sheet, data_workbook, data_sheet, output_table]


def check_equal_columns(sh_sheet, ds_sheet):
    for shell_row in sh_sheet.iter_rows(min_row=1, min_col=2):
        for data_row in ds_sheet.iter_rows(min_row=1, min_col=2):
            if len(shell_row) == len(data_row):
                return True
            else:
                return False


def insert_column(min_row, min_col, start_insert, result_sheet):
    i = start_insert
    for index, col in enumerate(result_sheet.iter_cols(min_row=min_row, min_col=min_col)):
        result_sheet.insert_cols(idx=i)
        prev_column_data = []
        for cell in col:
            prev_column_data.append(cell.value)

        for x in range(len(prev_column_data)):
            cell_to_write = result_sheet.cell(row=x + 1, column=i)
            cell_to_write.value = prev_column_data[x]
        i += 2
    return True


# function that handles unmerging cells in result shells that have mean(sd) merged for example.
def handle_merged_cells(result_sheet):
    merged_cells = result_sheet.merged_cells.ranges
    merged_cells_list = [item.coord for item in merged_cells]

    for merge_cell in merged_cells_list:
        left_cell_value = result_sheet[merge_cell.split(":")[0]].value
        result_sheet.unmerge_cells(merge_cell)
        result_sheet[merge_cell.split(":")[1]] = left_cell_value #copy value in left cell to right cell
    return True


def paste_data(sh_sheet, ds_sheet, st_row):
    data_loop_counter = 0
    for shell_row in sh_sheet.iter_rows(min_row=st_row, min_col=1):
        shell_row_values = tuple(shell_cell.value for shell_cell in shell_row)
        if any(map(lambda x: x is None, shell_row_values)):
            continue

        # #grab the last coordinate with data
        last_column_coord = shell_row[-1].coordinate
        for data_index, data_row in enumerate(ds_sheet.iter_rows(min_row=2, min_col=2)):
            if data_index == data_loop_counter:
                for index, cell in enumerate(shell_row[1:]):
                    sh_sheet[cell.coordinate] = data_row[index].value
                data_loop_counter += 1
                break
    return last_column_coord


def fetch_highlighted_columns(result_sheet, last_column_letter):
    section_with_colors = []
    for shell_col in result_sheet.iter_cols(min_row=1, min_col=1, max_col=1):
        colored_range = []
        for shell_col_index, shell_col_cell in enumerate(shell_col):
            if shell_col_cell.fill.start_color.index == "FFFFFF00":
                if shell_col[shell_col_index - 1].fill.start_color.index != "FFFFFF00":
                    colored_range.append(shell_col_cell.coordinate)
                elif shell_col[shell_col_index + 1].fill.start_color.index != "FFFFFF00":
                    colored_range.append(shell_col_cell.coordinate)
            else:
                if len(colored_range) > 0:
                    section_with_colors.append(colored_range)
                    colored_range = []

    # reassign the last coordinate of the highlighted sections to be the last column letter in table with data so it
    # know the range to sort
    if len(section_with_colors) > 0:
        for i in range(len(section_with_colors)):
            section_with_colors[i][-1] = f"{last_column_letter[0]}{section_with_colors[i][-1][1:]}"  #TODO: For last columns letter, think of how else to grab that info, and handle indexerror exception

        # create string for range of columns to be sorted
        for i in range(len(section_with_colors)):
            section_with_colors[i] = f"{section_with_colors[i][0]}:{section_with_colors[i][-1]}"
    return section_with_colors


def sort_rows_data(result_sheet, range_list):
    for ranges in range_list:
        cells_in_wb = result_sheet[ranges]

        cell_value_in_wb = []
        for cell_row in cells_in_wb:
            cell_list = []
            for each_cell in cell_row:
                # print(dir(each_cell))
                cell_list.append(each_cell.value)
            cell_list = tuple(cell_list)
            cell_value_in_wb.append(cell_list)

        cell_value_in_wb.sort(key=lambda a: a[1], reverse=True)  #TODO: include indexerror exception

        if len(cells_in_wb) > 0:
            for i, val1 in enumerate(cells_in_wb):
                for j, val2 in enumerate(val1):
                    cell_value_coordinate = val2.coordinate
                    result_sheet[cell_value_coordinate] = cell_value_in_wb[i][j]


def format_result_sheet(sh_sheet, col_inserted, st_row):
    count_loop = 0
    for row in sh_sheet.iter_rows(min_row=st_row, min_col=2):
        row_values = tuple(row_cell.value for row_cell in row)
        if any(map(lambda x: x is None, row_values)):
            continue

        if count_loop == 0:
            count_loop = 1
            first_cell = row[0].coordinate
            label_value = fetch_label(first_cell, sh_sheet)
            if label_value is None:
                continue

        label_value = fetch_label(row[0].coordinate, sh_sheet)
        grouped_cells = combine_cells(row)

        for cell in grouped_cells:
            first_cell, second_cell, first_value, second_value = split_cells(cell, sh_sheet)
            if not col_inserted:
                if "mean" in label_value.lower():
                    sh_sheet.merge_cells(cell)
                    sh_sheet[cell.split(":")[0]] = f"{first_value}({second_value})"
                elif any(value in label_value.lower() for value in ("95% ci", "q1, q3", "minimum, maximum", "min, max")):
                    sh_sheet.merge_cells(cell)
                    sh_sheet[cell.split(":")[0]] = f"({first_value}, {second_value})"
                elif any(value in label_value.lower() for value in ("median", "total patient sample", "number of patients")): #TODO: Logic faulty, if 'patients' in string of text line will run.
                    sh_sheet.merge_cells(cell)
                    sh_sheet[cell.split(":")[0]] = f"{first_value}"
                elif any(value in label_value.lower() for value in ("12", "24", "36", "48", "60")):
                    sh_sheet[cell.split(":")[0]] = f"{first_value}%"
                    sh_sheet[cell.split(":")[1]] = f"{second_value}"
                else:
                    sh_sheet[cell.split(":")[0]] = f"{first_value}"
                    sh_sheet[cell.split(":")[1]] = f"{second_value}%"
            else:
                if "mean" in label_value.lower():
                    sh_sheet[cell.split(":")[0]] = f"{first_value}({second_value})"
                elif any(value in label_value.lower() for value in ("95% ci", "q1, q3", "minimum, maximum", "min, max")):
                    sh_sheet[cell.split(":")[0]] = f"({first_value}, {second_value})"
                elif any(value in label_value.lower() for value in ("median", "total patient sample", "number of patients")): #TODO: Logic faulty, if 'patients' in string of text, line will run.
                    sh_sheet[cell.split(":")[0]] = f"{first_value}"
                elif any(value in label_value.lower() for value in ("12", "24", "36", "48", "60")):
                    sh_sheet[cell.split(":")[0]] = f"{first_value}%({second_value})"
                else:
                    sh_sheet[cell.split(":")[0]] = f"{first_value}({second_value}%)"


# This function is used to get the cell coordinates to place the count and the %, or whatever the second cell value
# should be.
def combine_cells(cell_list):
    count = 0
    cell_coordinate = []
    while count < len(cell_list) - 1:
        cell_group = f"{cell_list[count].coordinate}:{cell_list[count + 1].coordinate}"
        cell_coordinate.append(cell_group)
        count += 2
    return cell_coordinate


def split_cells(cell_value, result_sheet):
    f_cell = result_sheet[cell_value.split(":")[0]]
    s_cell = result_sheet[cell_value.split(":")[1]]
    f_value = f_cell.value
    s_value = s_cell.value
    return [f_cell, s_cell, f_value, s_value]


def fetch_label(current_cell, sh_sheet):
    adjacent_cell_value = sh_sheet[current_cell].offset(row=0, column=-1).value
    return adjacent_cell_value


def delete_column(result_sheet):
    m = 3
    for col in result_sheet.iter_cols(min_row=1, min_col=2):
        result_sheet.delete_cols(idx=m)
        m += 1


def main():
    pass


if __name__ == "__main__":
    main()