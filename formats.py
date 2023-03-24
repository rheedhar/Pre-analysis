from openpyxl import load_workbook
import time
import os
import re
from dotenv import load_dotenv

with open("00_formats.sas", "w") as format_file:
    format_file.write("proc format;\n")

wb = load_workbook("datamap.xlsx", data_only=True)
ws = wb["Sheet1"]
data = {}
i = 0
for row in ws.iter_rows(min_row=2, min_col=2):

    variable_name = row[0].value
    type_variable = row[2].value
    min_num, max_num = (row[3].value, row[4].value)

    if type_variable.lower() == "single answer" and min_num is None and max_num is None:
        result = {"variables": []}
        for item in row[5:]:
            if item.value is None:
                continue
            key_format = ws.cell(row=1, column=ws[item.coordinate].column).value
            value_format = item.value
            result[key_format] = value_format
        result["variables"].append(variable_name)

        if not data:
            data[f"format_{i}"] = result
            continue

        dict_equal = False
        for key, value in data.items():
            result_items = tuple((key, value) for key, value in result.items() if key != "variables")
            value_items = tuple((key, value) for key, value in value.items() if key != "variables")
            if result_items == value_items:
                data[key]["variables"].append(variable_name)
                dict_equal = True
                break
        if not dict_equal:
            i = i + 1
            data[f"format_{i}"] = result


for key, value in data.items():
    with open("00_formats.sas", "a", encoding="utf-8") as format_file1:
        format_file1.write(
            f"value {key}\n"
        )
    for key1, value1 in value.items():
        if key1 == "variables":
            continue

        with open("00_formats.sas", "a", encoding="utf-8") as format_file2:
            format_file2.write(
                f"{key1.split(' ')[1]}='{value1}'\n"
            )
    with open("00_formats.sas", "a", encoding="utf-8") as format_file3:
        format_file3.write(";\n")


with open("00_formats.sas", "a", encoding="utf-8") as format_file4:
    format_file4.write(
        "data a;\n set a;\n format \n "
    )

for key2, value2 in data.items():
    for var_value in value2["variables"]:
        with open("00_formats.sas", "a", encoding="utf-8") as format_file5:
            format_file5.write(f"{var_value}\n")
    with open("00_formats.sas", "a", encoding="utf-8") as format_file6:
        format_file6.write(f"{key2}.\n")

with open("00_formats.sas", "a", encoding="utf-8") as format_file6:
    format_file6.write("run;\n")