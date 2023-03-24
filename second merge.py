from openpyxl import load_workbook

name_workbook = input("Please enter the name of the excel workbook containing the result sheet without its extension ")
name_sheet = input("Please enter the name of the result sheet within the workbook ")
output_table = input("Please enter the name you want to call the output table without its extension ")
data_file = input("Please enter the name of the file containing the result data without its extension ")


main_workbook = load_workbook(f"data/{name_workbook}.xlsx")
sheets = main_workbook.sheetnames

if len(sheets) > 1:
    for s in sheets:
        if s != name_sheet:
            sheet_name = main_workbook.get_sheet_by_name(s)
            main_workbook.remove_sheet(sheet_name)

main_workbook.save(f"data/{output_table}.xlsx")
result_sheet = main_workbook.active


def insert_column():
    i = 3
    for index, col in enumerate(result_sheet.iter_cols(min_row=1, min_col=2)):
        result_sheet.insert_cols(idx=i)
        prev_column_data = []
        for cell in col:
            prev_column_data.append(cell.value)

        for x in range(len(prev_column_data)):
            cell_to_write = result_sheet.cell(row=x + 1, column=i)
            cell_to_write.value = prev_column_data[x]
        i += 2


def fetch_label(current_cell):
    adjacent_cell_value = updated_ws[current_cell].offset(row=0, column=-1).value
    return adjacent_cell_value


def combine_cells(cell_list):
    count = 0
    cell_coordinate = []
    while count < len(cell_list) - 1:
        cell_group = f"{cell_list[count].coordinate}:{cell_list[count + 1].coordinate}"
        cell_coordinate.append(cell_group)
        count += 2
    return cell_coordinate


def split_cells(cell_value):
    f_cell = updated_ws[cell_value.split(":")[0]]
    s_cell = updated_ws[cell_value.split(":")[1]]
    f_value = f_cell.value
    s_value = s_cell.value
    return [f_cell, s_cell, f_value, s_value]


insert_column()
main_workbook.save(f"data/{output_table}.xlsx")


# # load the new workbook(ie the result file) and the sheet you want to work on.
updated_wb = load_workbook(f"data/{output_table}.xlsx")
updated_ws = updated_wb[name_sheet]

# load the data workbook and the sheet you want to work on.
data_wb = load_workbook(f"data/{data_file}.xlsx")
data_ws = data_wb.active

# need to figure out what row to start pasting from
j = 0
last_column_coord = ""
first_time_loop_none = 0
for row in updated_ws.iter_rows(min_row=1, min_col=2):
    row_values = tuple(row_cell.value for row_cell in row)
    if any(map(lambda x: x is None, row_values)):
        continue

    if first_time_loop_none == 0:
        first_time_loop_none = 1

    if first_time_loop_none == 1:
        first_cell = row[0].coordinate
        label_value = fetch_label(first_cell)
        if label_value is None:
            continue

    # #grab the last coordinate with data
    last_column_coord = row[-1].coordinate
    for index2, row2 in enumerate(data_ws.iter_rows(min_row=2, min_col=2)):
        if index2 == j:
            for index, cell in enumerate(row):
                updated_ws[cell.coordinate] = row2[index].value
            j += 1
            break

# need to sort the data by highlighted columns.
# first we need to determine rows that have a background color.
list_section_bc = []
for col in updated_ws.iter_cols(min_row=1, min_col=1, max_col=1):
    colored_range = []

    for bracket in col:
        if bracket.fill.start_color.index == "FFFFFF00":
            colored_range.append(bracket.coordinate)
        else:
            if len(colored_range) > 0:
                list_section_bc.append(colored_range)
                colored_range = []

# reassign the last coordinate of the highlighted sections to be the last column in dataset.
for i in range(len(list_section_bc)):
    list_section_bc[i][-1] = f"{last_column_coord[0]}{list_section_bc[i][-1][1:]}"

# created final range of highlighted sections.
list_range_bc = []
for i in range(len(list_section_bc)):
    first_cell_coord = list_section_bc[i][0]
    last_cell_coord = list_section_bc[i][-1]
    final_range = f"{first_cell_coord}:{last_cell_coord}"
    list_range_bc.append(final_range)


# Fetching the values and putting them in a tuple of list so we can sort them.
for ranges in list_range_bc:
    cells_in_wb = updated_ws[ranges]

    cell_value_in_wb = []
    for cell_row in cells_in_wb:
        cell_list = []
        for each_cell in cell_row:
            # print(dir(each_cell))
            cell_list.append(each_cell.value)
        cell_list = tuple(cell_list)
        cell_value_in_wb.append(cell_list)

    cell_value_in_wb.sort(key=lambda a: a[1], reverse=True)

    if len(cells_in_wb) > 0:
        for i, val1 in enumerate(cells_in_wb):
            for j, val2 in enumerate(val1):
                cell_value_coordinate = val2.coordinate
                updated_ws[cell_value_coordinate] = cell_value_in_wb[i][j]

# # This for loop loops through the result file from row 3 and column 2. if any of the cells in a row is empty,
# # we continue to the next row. We basically want to format our values to look the way that we want it to look. so we
# # get the label of the first cell of the current row. Then we combine the 2 cells representing data for a column(ie
# # n, %) Then we loop through the grouped cell. but then we unpack each grouped cell values in separate variables. if
# # the label of the current row has a mean, then we merge the cell that we just unmerged and then format the value of
# # the merged cell to how we want it. we do the same for 95%, median, and otherwise just have n and % in separate
# # cells. Then we save the excel sheet.
#

second_time_loop_none = 0
for row in updated_ws.iter_rows(min_row=1, min_col=2):
    row_values = tuple(row_cell.value for row_cell in row)
    if any(map(lambda x: x is None, row_values)):
        continue

    if second_time_loop_none == 0:
        second_time_loop_none = 1

    if second_time_loop_none == 1:
        first_cell = row[0].coordinate
        label_value = fetch_label(first_cell)
        if label_value is None:
            continue

    first_cell = row[0].coordinate
    label_value = str(fetch_label(first_cell))
    grouped_cells = combine_cells(row)

    for cell in grouped_cells:
        first_cell, second_cell, first_value, second_value = split_cells(cell)
        if "mean" in label_value.lower():
            updated_ws[cell.split(":")[0]] = f"{first_value}({second_value})"
        elif any(value in label_value.lower() for value in ("95% ci", "q1, q3", "minimum, maximum", "min, max")):
            updated_ws[cell.split(":")[0]] = f"({first_value}, {second_value})"
        elif any(value in label_value.lower() for value in ("median", "total patient sample", "number of patients")):
            updated_ws[cell.split(":")[0]] = f"{first_value}"
        elif any(value in label_value.lower() for value in ("12", "24", "36", "48", "60")):
            updated_ws[cell.split(":")[0]] = f"{first_value}%({second_value})"
        else:
            updated_ws[cell.split(":")[0]] = f"{first_value}({second_value}%)"


def delete_column():
    m = 3
    for col in updated_ws.iter_cols(min_row=1, min_col=2):
        updated_ws.delete_cols(idx=m)
        m += 1


delete_column()

updated_wb.save(f"data/{output_table}.xlsx")

