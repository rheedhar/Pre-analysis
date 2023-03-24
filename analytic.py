from openpyxl import load_workbook
import time
import openai
import os
import re
from dotenv import load_dotenv


data = ["data a; \n", "set a; \n", "label \n"]
with open("labels.sas", "w") as label_file:
    label_file.writelines(data)

# API
load_dotenv()
openai_endpoint = "https://api.openai.com/v1/completions"
openai.api_key = os.getenv("OPEN_API_KEY")

wb = load_workbook("datamap.xlsx", data_only=True)
ws = wb["Sheet1"]


def handle_merged_cells():
    # TODO: We need to figure out how to detect that a cell is merged. Then we need to unmerge it and distribute the
    #  text in the left cell to their respective rows. We do this to the entire document first and then save the
    #  document.

    merged_cells = ws.merged_cells.ranges
    merged_cells_list_string = [item.coord for item in merged_cells]

    for merge_cell in merged_cells_list_string:
        merge_cell_range = merge_cell.split(":")
        left_cell_value = ws[merge_cell_range[0]].value  # left upper cell value.
        if "." in left_cell_value[:6]:
            if " " in left_cell_value[:6]:
                left_cell_value = left_cell_value.split(" ", 1)[1:] #we don't want the value before the full stop
            else:
                left_cell_value = left_cell_value.split(".", 1)[1:] #we don't want the value before the space.
            cell_value_item = left_cell_value[0].split("?")  # split items first by question mark
            main_label = cell_value_item[0]
            sub_labels = re.findall(r"\n\d+.+|\n.\t.+", cell_value_item[-1])
        else:
            cell_value_item = left_cell_value.split("\n", 1)
            main_label = cell_value_item[0]
            sub_labels = cell_value_item[1:][0].split("\n")

        sub_labels_stripped = [labels.strip().replace("\t", "") for labels in sub_labels]
        ws.unmerge_cells(merge_cell)  # unmerge the cells
        cell_range = ws[f"{merge_cell_range[0]}:{merge_cell_range[1]}"]  # get the cells in the workbook that were merged ws[c22:c30]

        for cell in cell_range:  # loop through the merged range
            # print(dir(cell_range[0][0]))
            current_cell = cell[0].coordinate  # get the coordinate for the current cell (the label)
            adjacent_cell_value = ws[current_cell].offset(row=0, column=-1).value  # get the adjacent cell(the variable name)
            if adjacent_cell_value.lower().endswith("dk"):
                ws[current_cell] = f"{main_label} - Don't Know"

            match = re.search(r"^.+_(\d+)_", adjacent_cell_value)
            if match:
                for text in sub_labels_stripped:
                    if text.startswith(match.group(1)):
                        sub_label = re.sub(r"\d+", "", text)
                        ws[current_cell] = f"{main_label} : {sub_label}"
                        if adjacent_cell_value.lower().endswith("dk"):
                            ws[current_cell] = f"{main_label} : {sub_label} - Don't Know"
                        break
            else:
                second_match = re.search(r"^([a-zA-Z0-9]+)[_|\.][a-zA-Z]+", adjacent_cell_value)
                if second_match:
                    for text in sub_labels_stripped:
                        if text.startswith(second_match.group(1)):
                            sub_label = text.split(".", 1)[1:]
                            ws[current_cell] = f"{main_label} : {sub_label}"
                            if adjacent_cell_value.lower().endswith("dk"):
                                ws[current_cell] = f"{main_label} : {sub_label} - Don't Know"
                            break
                else:
                    ws[current_cell] = f"{adjacent_cell_value}: {main_label}"
                    if adjacent_cell_value.lower().endswith("dk"):
                        ws[current_cell] = f"{adjacent_cell_value}: {main_label} - Don't Know"


handle_merged_cells()
wb.save("test1.xlsx")


wb1 = load_workbook("test1.xlsx", data_only=True)
ws1 = wb1["Sheet1"]
# TODO: Next we need to loop through each row in the worksheet, get the variable name and label, pass label to
#  chat gpt api and then save our result to our sas file.

for row in ws1.iter_rows(min_row=2, min_col=2, max_col=3, values_only=True):
    title, label = row
    # exclude_list = ["Patient_ID", "Abstraction_Date"]
    # if any([text in row for text in exclude_list]):
    #     new_label = label
    # else:
    #     new_label = label.split(" ", 1)[1]

    response = openai.Completion.create(
        model="text-davinci-003",
        prompt=f"The following text is a variable question: '${label}.' {'Create a descriptive label for the set of text before the : to describe the question being asked. Append the text after : to the label you create' if ':' in label else 'Create a descriptive label for the question being asked.'}",
        temperature=0
    )
    ai_label = (response["choices"][0]["text"]).strip()
    with open("labels.sas", "a", encoding="utf-8") as label_file:
        label_file.write(f"{title}='{ai_label}' \n")

with open("labels.sas", "a") as label_file:
    label_file.write("; \n run;")