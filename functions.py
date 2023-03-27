from openpyxl import load_workbook
import sys


def create_output_sheet():
    # Collect input data from user
    name_workbook = input(
        "Please enter the name of the excel workbook containing the result sheet without its extension ")
    name_sheet = input("Please enter the name of the result sheet within the workbook ")
    output_table = input("Please enter the name you want to call the output table without its extension ")
    data_file = input("Please enter the name of the file containing the result data without its extension ")

    if "" in (name_workbook, name_sheet, output_table, data_file):
        sys.exit("Please ensure to enter a file name")

    main_workbook = load_workbook(f"data/{name_workbook}.xlsx")
    sheets = main_workbook.sheetnames

    if len(sheets) > 1:
        for s in sheets:
            if s != name_sheet:
                sheet_name = main_workbook.get_sheet_by_name(s)
                main_workbook.remove_sheet(sheet_name)

    main_workbook.save(f"data/{output_table}.xlsx")
    result_sheet = main_workbook.active

    data_workbook = load_workbook(f"data/{data_file}.xlsx")
    data_sheet = data_workbook.active
    return [main_workbook, result_sheet, data_workbook, data_sheet, output_table]


# function that handles unmerging cells in result shells that have mean(sd) merged for example.
def handle_merged_cells(result_sheet):
    merged_cells = result_sheet.merged_cells.ranges
    if len(merged_cells) > 0:
        sheet_has_merged_cells = True
        merged_cells_list = [item.coord for item in merged_cells]
        for merge_cell in merged_cells_list:
            left_cell_value = result_sheet[merge_cell.split(":")[0]].value
            result_sheet.unmerge_cells(merge_cell)
            result_sheet[merge_cell.split(":")[1]] = left_cell_value
        return sheet_has_merged_cells


# function that inserts an extra column if the total column in data file is not the same length as the result file.
def insert_column(min_row, min_col, start_insert, result_sheet):
    i = start_insert
    for index, col in enumerate(result_sheet.iter_cols(min_row=min_row, min_col=min_col)):
        result_sheet.insert_cols(idx=i)
        prev_column_data = []
        for cell in col:
            prev_column_data.append(cell.value)

        for x in range(len(prev_column_data)):
            cell_to_write = result_sheet.cell(row=x + 1, column=i)
            cell_to_write.value = prev_column_data[x]
        i += 1
    return True


# This function is used to get the cell coordinates to place the count and the %, or whatever the second cell value
# should be.
def combine_cells(cell_list):
    count = 0
    cell_coordinate = []
    while count < len(cell_list) - 1:
        cell_group = f"{cell_list[count].coordinate}:{cell_list[count + 1].coordinate}"
        cell_coordinate.append(cell_group)
        count += 2
    return cell_coordinate


# in this function, we are trying to fetch the value of the adjacent cell(ie the label) of the current cell we are on.
# we used this mainly for the header row that we were trying to merge.
def fetch_label(current_cell, result_sheet):
    adjacent_cell_value = result_sheet[current_cell].offset(row=0, column=-1).value
    return adjacent_cell_value


def split_cells(cell_value, result_sheet):
    f_cell = result_sheet[cell_value.split(":")[0]]
    s_cell = result_sheet[cell_value.split(":")[1]]
    f_value = f_cell.value
    s_value = s_cell.value
    return [f_cell, s_cell, f_value, s_value]


def fetch_highlighted_columns(result_sheet, last_column_letter):
    list_section_bc = []
    for col in result_sheet.iter_cols(min_row=1, min_col=1, max_col=1):
        colored_range = []

        for index, bracket in enumerate(col):
            if bracket.fill.start_color.index == "FFFFFF00":
                if col[index - 1].fill.start_color.index != "FFFFFF00":
                    colored_range.append(bracket.coordinate)
                elif col[index + 1].fill.start_color.index != "FFFFFF00":
                    colored_range.append(bracket.coordinate)
            else:
                if len(colored_range) > 0:
                    list_section_bc.append(colored_range)
                    colored_range = []

    # reassign the last coordinate of the highlighted sections to be the last column letter in table with data so it
    # know the range to sort
    for i in range(len(list_section_bc)):
        list_section_bc[i][-1] = f"{last_column_letter[0]}{list_section_bc[i][-1][1:]}"

    # create string for range of columns to be sorted
    for i in range(len(list_section_bc)):
        list_section_bc[i] = f"{list_section_bc[i][0]}:{list_section_bc[i][-1]}"
    return list_section_bc


def sort_rows_data(result_sheet, range_list):
    for ranges in range_list:
        cells_in_wb = result_sheet[ranges]

        cell_value_in_wb = []
        for cell_row in cells_in_wb:
            cell_list = []
            for each_cell in cell_row:
                # print(dir(each_cell))
                cell_list.append(each_cell.value)
            cell_list = tuple(cell_list)
            cell_value_in_wb.append(cell_list)

        cell_value_in_wb.sort(key=lambda a: a[1], reverse=True)

        if len(cells_in_wb) > 0:
            for i, val1 in enumerate(cells_in_wb):
                for j, val2 in enumerate(val1):
                    cell_value_coordinate = val2.coordinate
                    result_sheet[cell_value_coordinate] = cell_value_in_wb[i][j]


def delete_column(result_sheet):
    m = 3
    for col in result_sheet.iter_cols(min_row=1, min_col=2):
        result_sheet.delete_cols(idx=m)
        m += 1