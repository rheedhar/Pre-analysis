from openpyxl import load_workbook
import sys


def extract_sheets():
    system_arguments = sys.argv

    if len(system_arguments) != 4:
        sys.exit("Please ensure to enter the correct number of file names")

    name_workbook, name_sheet, format_file = system_arguments[1:]

    # if any of the variables is empty, exit from the entire program
    if " " in (name_workbook, name_sheet, format_file):
        sys.exit("Please ensure to enter a file name")

    workbook = load_workbook(f"data/{name_workbook}", data_only=True)
    worksheet = workbook[name_sheet]

    return workbook, worksheet, format_file


def create_data_struct(sh_sheet):
    # create empty dictionary
    data = {}
    format_counter = 0
    for row in sh_sheet.iter_rows(min_row=2, min_col=2):

        variable_name = row[0].value
        type_variable = row[2].value
        min_num, max_num = (row[3].value, row[4].value)

        if type_variable.lower() == "single answer" and min_num is None and max_num is None:
            current_format = {"variables": []}
            for cell in row[5:]:
                if cell.value is None:
                    continue
                key_format = sh_sheet.cell(row=1, column=sh_sheet[cell.coordinate].column).value  # grab the key from the header
                value_format = cell.value
                current_format[key_format] = value_format
            current_format["variables"].append(variable_name)

            if not data:
                data[f"format_{format_counter}"] = current_format
                continue

            dict_equal = False
            for key, value in data.items():
                result_items = tuple((key, value) for key, value in current_format.items() if key != "variables")
                value_items = tuple((key, value) for key, value in value.items() if key != "variables")

                if result_items == value_items:
                    data[key]["variables"].append(variable_name)
                    dict_equal = True
                    break
            if not dict_equal:
                format_counter = format_counter + 1
                data[f"format_{format_counter}"] = current_format

    return data


def create_format_file(data_file, file_name):

    with open(f"data/{file_name}", "w") as format_file:
        format_file.write("proc format;\n")

    for key, value in data_file.items():
        with open(f"data/{file_name}", "a", encoding="utf-8") as format_file1:
            format_file1.write(
                f"value {key}\n"
            )
        for key1, value1 in value.items():
            if key1 == "variables":
                continue

            with open(f"data/{file_name}", "a", encoding="utf-8") as format_file2:
                format_file2.write(
                    f"{key1.split(' ')[1]}='{value1}'\n"
                )
        with open(f"data/{file_name}", "a", encoding="utf-8") as format_file3:
            format_file3.write(";\n")

    with open(f"data/{file_name}", "a", encoding="utf-8") as format_file4:
        format_file4.write(
            "data a;\n set a;\n format \n "
        )

    for key2, value2 in data_file.items():
        for var_value in value2["variables"]:
            with open(f"data/{file_name}", "a", encoding="utf-8") as format_file5:
                format_file5.write(f"{var_value}\n")
        with open(f"data/{file_name}", "a", encoding="utf-8") as format_file6:
            format_file6.write(f"{key2}.\n")

    with open(f"data/{file_name}", "a", encoding="utf-8") as format_file6:
        format_file6.write("run;\n")


def main():
    pass


if __name__ == "__main__":
    main()